import argparse, shutil, csv, re
from pathlib import Path
import json
import subprocess
from collections import defaultdict
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import socket

from tests.SAVR2SAVR7 import SAVR7 #01
from tests.SAVR2SAVR14 import SAVR14 #08
from tests.SAVR2SAVR13 import SAVR13 #09
from tests.SAVR2SAVR18 import SAVR18 #05
from tests.SAVR2SAVR15 import SAVR15 #07
from tests.SAVR2SAVR43 import SAVR43_1, SAVR43_2, SAVR43_3
from tests.SAVR2SAVR16 import SAVR16
from tests.SAVR2SAVR6 import SAVR6
from tests.SAVR9 import SAVR9
from tests.SAVR17 import SAVR17
from tests.SAVR4 import SAVR4
from tests.SAVR29 import SAVR29
from tests.SAVR12 import SAVR12
from tests.SAVR27a28 import SAVR27a28
from tests.SAVR40 import SAVR40
from tests.SAVR5 import SAVR5
from tests.SAVR44 import SAVR44
from tests.SAVR41 import SAVR41

#python overall.py --start "2026-07-02 16:00:00.049" --roster roster.json --out results.csv

TEST_CLASSES = [SAVR4, SAVR5, SAVR6, SAVR7, SAVR9, SAVR12, SAVR13, SAVR14, SAVR15, 
SAVR16, SAVR17, SAVR18, SAVR27a28, SAVR29, SAVR40, SAVR41, SAVR43_1, SAVR43_2, SAVR43_3, SAVR44]

LOG_PATH    = Path(r"C:\Windows\System32\config\systemprofile\AppData\Local\Cybersenz\SecureAiService\Logs\SecureAiService.log")
AGENTS_PATH = Path(r"C:\ProgramData\Cybersenz\config\agents\detected_agents.json")
SYSINFO_PATH = Path(r"C:\ProgramData\Cybersenz\config\sysinfo.jsonl")
SESSION_TOKEN = ""

def login_dev():
    #Login using dev QA credentials
    url = "https://api-dev.cybersenz.com/api/v1/auth/login"
    headers = {"email": "qa", "password": "Qa"}
    #email and password to either be in a .env or entered by user. Change this to reflect that.
    response = requests.get(url, headers=headers)
    data = response.json()

    #Check if successful
    if not data["success"]:
        print("Credentials not working. API QA terminated.")
        return false
    
    #Verify OTP
    url = "https://api-dev.cybersenz.com/api/v1/auth/verify-otp"
    headers = {"email": "qa", "otp": "000000"}
    #OTP to be in a .env or entered by user. Change this to reflect that.
    response = requests.get(url, headers=headers)
    data = response.json()

    #Check if successful
    if not data["success"]:
        print("OTP stage not working. API QA terminated.")
        return false
    
    #Extract Session Token
    SESSION_TOKEN = data["session_token"]

def get_API():
    #Create header using bearer token for rest of urls
    headers = {
    "Authorization": "Bearer " + SESSION_TOKEN
    }
    device_name = socket.gethostname() #Get Device Name

    #Get list of devices
    url = "https://api-dev.cybersenz.com/api/v1/devices/"
    response = requests.get(url, headers=headers)
    data = response.json()

def get_pids_by_name(proc_name):
    result = subprocess.run(
        ["tasklist", "/FI", f"IMAGENAME eq {proc_name}", "/FO", "CSV", "/NH"],
        capture_output=True, text=True
    )
    pids = []
    for line in result.stdout.strip().splitlines():
        parts = line.strip('"').split('","')
        if len(parts) >= 2:
            pids.append(parts[1])
    return pids

def build_tests(roster_path, agents, sysinfo, test_filter=None):
    if roster_path is None:
        return []
    cfg = json.loads(roster_path.read_text())

    # auto-resolve tcp_stats_test by_name entries into by_pid using live tasklist
    if "SAVR14" in cfg:
        tcp_cfg = cfg["SAVR14"]
        for proc_name, domain in tcp_cfg.get("by_name", {}).items():
            pids = get_pids_by_name(proc_name)
            print(f"looked up {proc_name} -> PIDs: {pids}")
            for pid in pids:
                tcp_cfg["by_pid"][pid] = {"label": proc_name, "domain": domain}
        tcp_cfg["by_name"] = {}

    allowed = {f"SAVR{n}" for n in test_filter} if test_filter else None

    return [cls(cfg[cls.name], agents, sysinfo)
            for cls in TEST_CLASSES
            if cls.name in cfg and (allowed is None or cls.name in allowed)]


def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--start", required=True,
                   help="Run-start timestamp, e.g. '2026-06-25 08:13:00.000'")
    p.add_argument("--out", default=Path("results.xlsx"), type=Path)
    p.add_argument("--roster", type=Path,
                   help="JSON of expected subjects -> conf range")
    p.add_argument("--tests", nargs="*", default=None,
                   metavar="SAVR",
                   help="Run only these tests, e.g. --tests 4 5 7 43_1")
    return p.parse_args()

def snapshot(log_path: Path) -> Path:
    # write locally, not next to the protected log file
    snap = Path("log.snapshot")
    shutil.copy2(log_path, snap)
    return snap

def snapshot_agents(agents_path: Path) -> Path:
    snap = Path("agents.snapshot")
    shutil.copy2(agents_path, snap)
    return snap

timestamp_re = re.compile(r"^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}\.\d{3})")

def load_window(snap: Path, start: str) -> list[str]:
    lines = snap.read_text(encoding="utf-8", errors="replace").splitlines()
    out, started = [], False
    for line in lines:
        m = timestamp_re.match(line)
        if not started:
            if m and m.group(1) >= start:   # >= not ==, string compare is safe for this format
                started = True
            else:
                continue
        out.append(line)
    return out   # indexable array; multi-line matchers can peek forward

def load_agents(agents_snap: Path, start: str) -> list[dict]:
    # parse start as local naive datetime then treat as UTC for comparison
    start_dt = datetime.strptime(start, "%Y-%m-%d %H:%M:%S.%f").replace(
        tzinfo=timezone.utc
    )
    data = json.loads(agents_snap.read_text(encoding="utf-8"))
    active = []
    for agent in data.get("agents", []):
        last_seen = agent.get("last_seen", "")
        if not last_seen:
            continue
        try:
            ls_dt = datetime.fromisoformat(last_seen.replace("Z", "+00:00"))
        except ValueError:
            continue
        if ls_dt >= start_dt:
            active.append(agent)
    return active

def load_sysinfo(sysinfo_path: Path) -> dict:
    with sysinfo_path.open("rb") as f:
        f.seek(0, 2)          # seek to end of file
        pos = f.tell()
        
        # walk backwards skipping any trailing newlines
        while pos > 0:
            pos -= 1
            f.seek(pos)
            if f.read(1) not in (b"\n", b"\r", b" "):
                break
        
        # now find the start of this last line
        while pos > 0:
            pos -= 1
            f.seek(pos)
            if f.read(1) in (b"\n", b"\r"):
                break
        
        last_line = f.readline().decode("utf-8").strip()
    
    return json.loads(last_line) if last_line else {}

def run(window, tests):
    for i, line in enumerate(window):
        for t in tests:
            t.offer(line, i, window)
    for t in tests:
        t.resolve()

# column widths (Excel character units) and wrap behavior per column
COLUMN_WIDTHS = {
    "test":     12,
    "subject":  28,
    "expected": 38,
    "actual":   42,
    "result":   14,
    "comments": 55,
}
RESULT_FILL = {
    "PASS":        "C6EFCE",
    "FAIL":        "FFC7CE",
    "NOT_DETECTED": "FFEB9C",
    "PARTIAL":     "FFEB9C",
    "INCONCLUSIVE": "D9D9D9",
}

def write_report(tests, out: Path):
    from openpyxl.styles import PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "results"

    headers = ["test", "subject", "expected", "actual", "result", "comments"]
    ws.append(headers)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS[header]

    row_idx = 2
    for t in tests:
        for row in t.rows():
            ws.append(row)
            result_val = row[4] if len(row) > 4 else ""
            fill_color = RESULT_FILL.get(str(result_val).strip())
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if fill_color and headers[col_idx - 1] == "result":
                    cell.fill = PatternFill(start_color=fill_color,
                                             end_color=fill_color,
                                             fill_type="solid")
            row_idx += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(out)

def summarize_report(tests):
    # Aggregate counts per test name
    counts = defaultdict(lambda: {"total": 0, "PASS": 0, "FAIL": 0, 
                                   "PARTIAL": 0, "NOT_DETECTED": 0, "INCONCLUSIVE": 0})
    for t in tests:
        name = t.name  # e.g. "SAVR7"
        for row in t.rows():
            # row = [test, subject, expected, actual, result, comments]
            result = row[4].strip() if len(row) > 4 else ""
            counts[name]["total"] += 1
            if result in counts[name]:
                counts[name][result] += 1

    # Print in SAVR-number order
    def savr_sort_key(name):
        m = re.search(r"(\d+)", name)
        return int(m.group(1)) if m else 0

    for name in sorted(counts, key=savr_sort_key):
        c = counts[name]
        print(
            f"[Test] {name}: {c['total']} items, "
            f"{c['PASS']} PASS, {c['FAIL']} FAIL, "
            f"{c['PARTIAL']} PARTIAL, {c['NOT_DETECTED']} NOT_DETECTED, "
            f"{c['INCONCLUSIVE']} INCONCLUSIVE"
        )

def main():
    a = parse_args()
    snap        = snapshot(LOG_PATH)
    agents_snap = snapshot_agents(AGENTS_PATH)
    window      = load_window(snap, a.start)
    agents      = load_agents(agents_snap, a.start)  # fixed: use load_agents not raw json.loads
    sysinfo = load_sysinfo(SYSINFO_PATH)
    tests = build_tests(a.roster, agents, sysinfo, a.tests)
    run(window, tests)

    #API Stage
    if (login_dev()):
        get_API()
    
    write_report(tests, a.out)
    summarize_report(tests)

if __name__ == "__main__":
    main()
